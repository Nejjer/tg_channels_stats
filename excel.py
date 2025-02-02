import random
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from channel_statistic import ChannelStatistic


def save_to_excel_with_formatting(stats: List[ChannelStatistic], filename: str):
    # Преобразуем список объектов в список словарей
    data = [stat.__dict__ for stat in stats]
    # Создаем DataFrame из списка словарей
    df = pd.DataFrame(data)
    # Сохраняем DataFrame в Excel файл
    df.to_excel(filename, index=False)

    # Загружаем Excel файл с помощью openpyxl
    wb = load_workbook(filename)
    ws = wb.active

    # Рассчитываем перцентили для всех числовых столбцов
    percentiles_75 = {}
    percentiles_90 = {}
    percentiles_10 = {}
    percentiles_25 = {}
    for column in df.columns:
        if pd.api.types.is_numeric_dtype(df[column]):
            percentiles_75[column] = df[column].quantile(0.75)
            percentiles_10[column] = df[column].quantile(0.10)
            percentiles_25[column] = df[column].quantile(0.25)
            percentiles_90[column] = df[column].quantile(0.90)

    # Задаем стили для заливки
    yellow_fill_25 = PatternFill(start_color="ffcf69", end_color="ffcf69", fill_type="solid")
    red_fill_10 = PatternFill(start_color="ff6969", end_color="ff6969", fill_type="solid")
    green_fill_90 = PatternFill(start_color="2ef520", end_color="2ef520", fill_type="solid")
    white_green_fill_75 = PatternFill(start_color="c0ff69", end_color="c0ff69", fill_type="solid")

    # Применяем форматирование
    for col_idx, column in enumerate(df.columns, start=1):  # Перебираем столбцы
        if column in percentiles_75 and column in percentiles_10:  # Проверяем только числовые столбцы
            perc_75 = percentiles_75[column]
            perc_10 = percentiles_10[column]
            perc_25 = percentiles_25[column]
            perc_90 = percentiles_90[column]
            for row_idx in range(2, len(df) + 2):  # Перебираем строки (пропуская заголовок)
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    if cell.value < perc_10:
                        cell.fill = red_fill_10
                    elif cell.value < perc_25:
                        cell.fill = yellow_fill_25
                    elif cell.value > perc_90:
                        cell.fill = green_fill_90
                    elif cell.value > perc_75:
                        cell.fill = white_green_fill_75

    # Делаем легенду для цветов
    cell = ws['K1']
    cell.value = '10% лучшего'
    cell.fill = green_fill_90
    cell = ws['K2']
    cell.value = '25% лучшего'
    cell.fill = white_green_fill_75
    cell = ws['K3']
    cell.value = '25% худшего'
    cell.fill = yellow_fill_25
    cell = ws['K4']
    cell.value = '10% худшего'
    cell.fill = red_fill_10
    # Устанавливаем автоматическую ширину столбцов
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Получаем букву столбца
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width
    # Сохраняем файл
    wb.save(filename)


# Пример использования
if __name__ == "__main__":
    stats = [
        ChannelStatistic(
            name=f"Channel{i + 1}",
            count_subs=random.randint(500, 5000),
            average_views=round(random.uniform(100, 2000), 1),
            average_reacts=round(random.uniform(10, 300), 1),
            freq_posts_per_week=round(random.uniform(1, 14), 1),
            average_comments=round(random.uniform(5, 100), 1),
            average_message_length=round(random.uniform(50, 500), 1),
            engagement_rate=round(random.uniform(0, 100), 1)
        ) for i in range(15)
    ]
    stats.append(ChannelStatistic('name', 1, 1, 1, 1, 1, 1, 1))

    save_to_excel_with_formatting(stats, "channel_statistics.xlsx")
    print("Данные успешно сохранены в channel_statistics.xlsx")
